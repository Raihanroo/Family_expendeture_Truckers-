"""
Family Expenditure Management System - Views
Handles all web views and API endpoints for expense tracking
"""
import json
import os
import openpyxl
from datetime import datetime, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from rest_framework import permissions, viewsets

from .forms import BudgetForm, ExpenseForm, FamilyMemberForm
from .models import (
    Budget,
    Expense,
    ExpenseCategory,
    Expenditure,
    FamilyMember,
    IncomeSource,
)
from .serializers import (
    ExpenseCategorySerializer,
    ExpenditureSerializer,
    ExpenseSerializer,
    FamilyMemberSerializer,
    IncomeSourceSerializer,
)



# --- API ViewSets ---
class IncomeSourceViewSet(viewsets.ModelViewSet):
    queryset = IncomeSource.objects.all()
    serializer_class = IncomeSourceSerializer
    permission_classes = [permissions.IsAuthenticated]


class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes = [permissions.IsAuthenticated]


class FamilyMemberViewSet(viewsets.ModelViewSet):
    serializer_class = FamilyMemberSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return FamilyMember.objects.filter(user=self.request.user)


class ExpenditureViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenditureSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Expenditure.objects.filter(member__user=self.request.user)


class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user).order_by("-date")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# --- Admin Dashboard ---
@login_required
def admin_dashboard(request):
    """Admin dashboard with system-wide statistics and charts"""
    if not request.user.is_superuser:
        messages.error(request, "Access denied. Admin only.")
        return redirect("expenses:home")

    today = datetime.today()
    chart_period = request.GET.get("period", "7days")  # Default to 7 days
    all_expenses = Expense.objects.all().select_related("category", "member__user")

    total_this_month = (
        all_expenses.filter(
            date__year=today.year, date__month=today.month
        ).aggregate(Sum("amount"))["amount__sum"]
        or 0
    )
    total_all_time = all_expenses.aggregate(Sum("amount"))["amount__sum"] or 0
    total_expenses_count = all_expenses.count()
    expenses_this_month = all_expenses.filter(
        date__year=today.year, date__month=today.month
    ).count()

    total_members = FamilyMember.objects.count()
    admin_count = FamilyMember.objects.filter(role="ADMIN").count()
    member_count = FamilyMember.objects.filter(role="MEMBER").count()
    total_categories = ExpenseCategory.objects.count()
    total_income_sources = IncomeSource.objects.count()

    monthly_budget = Budget.objects.filter(user=request.user).first()
    budget_amount = monthly_budget.monthly_budget if monthly_budget else 0
    budget_percentage = (
        (total_this_month / budget_amount * 100) if budget_amount > 0 else 0
    )

    current_month_expenses = all_expenses.filter(
        date__year=today.year, date__month=today.month
    )
    category_breakdown = (
        current_month_expenses.values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )
    pie_labels = [item["category__name"] or "General" for item in category_breakdown]
    pie_data = [float(item["total"]) for item in category_breakdown]

    # Dynamic chart period calculation
    if chart_period == "7days":
        start_date = today - timedelta(days=7)
        date_format = "%d %b"
        chart_title = "Last 7 Days"
    elif chart_period == "30days":
        start_date = today - timedelta(days=30)
        date_format = "%d %b"
        chart_title = "Last 30 Days"
    elif chart_period == "3months":
        start_date = today - timedelta(days=90)
        date_format = "%d %b"
        chart_title = "Last 3 Months"
    elif chart_period == "6months":
        start_date = today - timedelta(days=180)
        date_format = "%b %Y"
        chart_title = "Last 6 Months"
    elif chart_period == "year":
        start_date = today - timedelta(days=365)
        date_format = "%b %Y"
        chart_title = "Last Year"
    else:
        start_date = today - timedelta(days=7)
        date_format = "%d %b"
        chart_title = "Last 7 Days"

    daily_spending = (
        all_expenses.filter(date__gte=start_date)
        .values("date")
        .annotate(total=Sum("amount"))
        .order_by("date")
    )
    bar_labels = [item["date"].strftime(date_format) for item in daily_spending]
    bar_data = [float(item["total"]) for item in daily_spending]

    context = {
        "total_this_month": total_this_month,
        "total_all_time": total_all_time,
        "total_expenses_count": total_expenses_count,
        "expenses_this_month": expenses_this_month,
        "total_members": total_members,
        "admin_count": admin_count,
        "member_count": member_count,
        "total_categories": total_categories,
        "total_income_sources": total_income_sources,
        "monthly_budget": budget_amount,
        "budget_percentage": budget_percentage,
        "recent_expenses": all_expenses.order_by("-date")[:10],
        "pie_labels": json.dumps(pie_labels),
        "pie_data": json.dumps(pie_data),
        "bar_labels": json.dumps(bar_labels),
        "bar_data": json.dumps(bar_data),
        "chart_period": chart_period,
        "chart_title": chart_title,
    }
    return render(request, "expenses/admin_dashboard.html", context)


# --- Home / Dashboard ---
@login_required
def home(request):
    """
    User dashboard showing only their own expenses.
    Admin users should use admin_dashboard for system-wide view.
    """
    user = request.user
    search_query = request.GET.get("search", "").strip()
    chart_period = request.GET.get("period", "7days")  # Default to 7 days

    # Each user sees only their own expenses
    all_expenses = Expense.objects.filter(user=user).select_related(
        "category", "member"
    ).order_by("-date")

    converted_date = None
    if search_query:
        for fmt in ("%d %b, %Y", "%d %b %Y", "%Y-%m-%d"):
            try:
                converted_date = datetime.strptime(search_query, fmt).strftime(
                    "%Y-%m-%d"
                )
                break
            except ValueError:
                continue

        query_filter = (
            Q(description__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(member__name__icontains=search_query)
        )
        if converted_date:
            query_filter |= Q(date=converted_date)
        expenses = all_expenses.filter(query_filter).distinct()
    else:
        expenses = all_expenses[:10]

    today = datetime.today()
    current_month_expenses = all_expenses.filter(
        date__year=today.year, date__month=today.month
    )
    total_this_month = (
        current_month_expenses.aggregate(Sum("amount"))["amount__sum"] or 0
    )
    total_all_time = all_expenses.aggregate(Sum("amount"))["amount__sum"] or 0

    category_breakdown = (
        current_month_expenses.values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )
    pie_labels = [item["category__name"] or "General" for item in category_breakdown]
    pie_data = [float(item["total"]) for item in category_breakdown]

    # Dynamic chart period calculation
    if chart_period == "7days":
        start_date = today - timedelta(days=7)
        date_format = "%d %b"
        chart_title = "Last 7 Days"
    elif chart_period == "30days":
        start_date = today - timedelta(days=30)
        date_format = "%d %b"
        chart_title = "Last 30 Days"
    elif chart_period == "3months":
        start_date = today - timedelta(days=90)
        date_format = "%d %b"
        chart_title = "Last 3 Months"
    elif chart_period == "6months":
        start_date = today - timedelta(days=180)
        date_format = "%b %Y"
        chart_title = "Last 6 Months"
    elif chart_period == "year":
        start_date = today - timedelta(days=365)
        date_format = "%b %Y"
        chart_title = "Last Year"
    else:
        start_date = today - timedelta(days=7)
        date_format = "%d %b"
        chart_title = "Last 7 Days"

    daily_spending = (
        all_expenses.filter(date__gte=start_date)
        .values("date")
        .annotate(total=Sum("amount"))
        .order_by("date")
    )
    bar_labels = [item["date"].strftime(date_format) for item in daily_spending]
    bar_data = [float(item["total"]) for item in daily_spending]

    context = {
        "expenses": expenses,
        "total_this_month": total_this_month,
        "total_all_time": total_all_time,
        "total_expenses_count": all_expenses.count(),
        "pie_labels": json.dumps(pie_labels),
        "pie_data": json.dumps(pie_data),
        "bar_labels": json.dumps(bar_labels),
        "bar_data": json.dumps(bar_data),
        "chart_period": chart_period,
        "chart_title": chart_title,
        "is_admin": user.is_superuser,
    }
    return render(request, "expenses/home.html", context)


# --- Export Excel ---
@login_required
def export_expenses_excel(request):
    search_query = request.GET.get("search", "").strip()
    expenses = Expense.objects.filter(user=request.user).select_related(
        "category", "member"
    ).order_by("-date")

    if search_query:
        query_filter = (
            Q(description__icontains=search_query)
            | Q(category__name__icontains=search_query)
            | Q(member__name__icontains=search_query)
            | Q(member__user__username__icontains=search_query)
        )
        expenses = expenses.filter(query_filter).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Category", "Member", "Description", "Amount"])
    for obj in expenses:
        ws.append([
            obj.date.strftime("%d %b, %Y"),
            obj.category.name if obj.category else "General",
            obj.member.name if obj.member else "Self",
            obj.description or "",
            float(obj.amount),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Report_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    )
    wb.save(response)
    return response


# --- Expense CRUD ---
@login_required
def add_expense(request):
    """Add a new expense - member dropdown shows only user's own members"""
    if request.method == "POST":
        form = ExpenseForm(request.POST, user=request.user)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            messages.success(request, "Expense added successfully!")
            return redirect("expenses:home")
    else:
        form = ExpenseForm(user=request.user)
    return render(request, "expenses/add_expense.html", {"form": form})


@login_required
def edit_expense(request, pk):
    """Edit an existing expense - users can only edit their own expenses"""
    try:
        expense = get_object_or_404(Expense, pk=pk, user=request.user)
    except:
        messages.error(request, "Expense not found or you don't have permission to edit it!")
        return redirect("expenses:home")
    
    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense updated successfully!")
            return redirect("expenses:home")
    else:
        form = ExpenseForm(instance=expense, user=request.user)
    return render(
        request, "expenses/add_expense.html", {"form": form, "edit_mode": True}
    )


@login_required
def delete_expense(request, pk):
    """Delete an expense - users can only delete their own expenses"""
    try:
        expense = get_object_or_404(Expense, pk=pk, user=request.user)
        expense.delete()
        messages.success(request, "Expense deleted successfully!")
    except:
        messages.error(request, "Expense not found or you don't have permission to delete it!")
    return redirect("expenses:home")


# --- Category Management ---
@login_required
def add_category(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            ExpenseCategory.objects.create(name=name)
            messages.success(request, "Category added!")
        return redirect("expenses:add_category")
    return render(
        request,
        "expenses/add_category.html",
        {"categories": ExpenseCategory.objects.all()},
    )


@login_required
def edit_category(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            category.name = name
            category.save()
            messages.success(request, "Category updated!")
        return redirect("expenses:add_category")
    return render(request, "expenses/edit_category.html", {"category": category})


@login_required
def delete_category(request, pk):
    get_object_or_404(ExpenseCategory, pk=pk).delete()
    return redirect("expenses:add_category")


# --- Member Management ---
@login_required
def add_member(request):
    if request.method == "POST":
        form = FamilyMemberForm(request.POST, request.FILES)
        if form.is_valid():
            member = form.save(commit=False)
            member.user = request.user
            member.save()
            messages.success(request, "Member added successfully!")
            return redirect("expenses:member_list")
    else:
        form = FamilyMemberForm()
    return render(
        request,
        "expenses/add_member.html",
        {"form": form, "roles": FamilyMember.ROLE_CHOICES},
    )


@login_required
def member_list(request):
    """List all family members - admins see all, regular users see only their own"""
    query = request.GET.get("q", "").strip()
    
    # Admin can see all members, regular users see only their own
    if request.user.is_superuser:
        members = FamilyMember.objects.all()
    else:
        members = FamilyMember.objects.filter(user=request.user)

    if query:
        members = members.filter(
            Q(name__icontains=query) |
            Q(phone_number__icontains=query) |
            Q(father_name__icontains=query) |
            Q(mother_name__icontains=query) |
            Q(address__icontains=query) |
            Q(income_source__icontains=query)
        )

    if request.GET.get("export") == "1":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Role", "Income Source", "Salary", "Owner"])
        for m in members:
            ws.append([
                m.name, 
                m.phone_number, 
                m.role, 
                m.income_source, 
                float(m.salary or 0),
                m.user.username
            ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=members.xlsx"
        wb.save(response)
        return response

    return render(request, "expenses/member_list.html", {
        "members": members, 
        "search_query": query,
        "is_admin": request.user.is_superuser
    })


@login_required
def edit_member(request, pk):
    """Edit family member - admins can edit any member, users can edit only their own"""
    # Admin can edit any member, regular users can edit only their own
    if request.user.is_superuser:
        member = get_object_or_404(FamilyMember, id=pk)
    else:
        member = get_object_or_404(FamilyMember, id=pk, user=request.user)
    
    if request.method == "POST":
        form = FamilyMemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, "Member updated successfully!")
            return redirect("expenses:member_list")
    else:
        form = FamilyMemberForm(instance=member)
    return render(
        request,
        "expenses/add_member.html",
        {"form": form, "member": member, "edit_mode": True, "roles": FamilyMember.ROLE_CHOICES},
    )


@login_required
def delete_member(request, pk):
    """Delete family member - admins can delete any member, users can delete only their own"""
    try:
        if request.user.is_superuser:
            member = get_object_or_404(FamilyMember, id=pk)
        else:
            member = get_object_or_404(FamilyMember, id=pk, user=request.user)
        member.delete()
        messages.success(request, "Member deleted successfully!")
    except:
        messages.error(request, "Member not found or you don't have permission!")
    return redirect("expenses:member_list")


# --- Authentication ---
def register_view(request):
    if request.method == "POST":
        u = request.POST.get("username")
        e = request.POST.get("email")
        p1 = request.POST.get("password")
        p2 = request.POST.get("password_confirm")
        if p1 == p2 and not User.objects.filter(username=u).exists():
            user = User.objects.create_user(username=u, email=e, password=p1)
            user.is_staff = user.is_superuser = True
            user.save()
            return redirect("expenses:login")
    return render(request, "register.html")


def login_view(request):
    if request.method == "POST":
        u = request.POST.get("username")
        p = request.POST.get("password")
        user = authenticate(username=u, password=p)
        if user:
            login(request, user)
            return redirect("expenses:home")
        messages.error(request, "Invalid username or password!")
        return render(request, "login.html", {"error": "Invalid username or password!"})
    return render(request, "login.html")


def admin_register(request):
    """Admin registration — limited to MAX_ADMINS superusers."""
    MAX_ADMINS = int(os.environ.get("MAX_ADMINS", 3))
    ADMIN_SECRET = os.environ.get("ADMIN_SECRET_KEY", "")

    if User.objects.filter(is_superuser=True).count() >= MAX_ADMINS:
        messages.error(request, f"Maximum admin limit ({MAX_ADMINS}) reached.")
        return redirect("expenses:login")

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        secret_key = request.POST.get("secret_key", "")

        if password != confirm_password:
            return render(request, "register.html", {"error": "Passwords do not match!"})
        if User.objects.filter(username=username).exists():
            return render(request, "register.html", {"error": "Username already exists!"})
        if User.objects.filter(email=email).exists():
            return render(request, "register.html", {"error": "Email already exists!"})
        if not ADMIN_SECRET or secret_key != ADMIN_SECRET:
            return render(request, "register.html", {"error": "Invalid secret key!"})

        User.objects.create_user(
            username=username,
            email=email,
            password=password,
            is_superuser=True,
            is_staff=True,
        )
        messages.success(request, "Admin account created! Please login.")
        return redirect("expenses:login")

    return render(request, "register.html")


def logout_view(request):
    logout(request)
    return redirect("expenses:login")


def admin_login_view(request):
    if request.method == "POST":
        u = request.POST.get("username")
        p = request.POST.get("password")
        user = authenticate(username=u, password=p)
        if user:
            if user.is_superuser:
                login(request, user)
                return redirect("expenses:admin_dashboard")
            return render(
                request,
                "login_admin.html",
                {"error": "Not an admin account! Use user login."},
            )
        return render(
            request, "login_admin.html", {"error": "Invalid username or password!"}
        )
    return render(request, "login_admin.html")


def user_register(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password != confirm_password:
            return render(request, "register_user.html", {"error": "Passwords do not match!"})
        if User.objects.filter(username=username).exists():
            return render(request, "register_user.html", {"error": "Username already exists!"})
        if User.objects.filter(email=email).exists():
            return render(request, "register_user.html", {"error": "Email already exists!"})

        User.objects.create_user(
            username=username,
            email=email,
            password=password,
            is_superuser=False,
            is_staff=False,
        )
        messages.success(request, "Account created successfully! Please login.")
        return redirect("expenses:login")

    return render(request, "register_user.html")


@login_required
def set_budget(request):
    budget, _ = Budget.objects.get_or_create(user=request.user)
    if request.method == "POST":
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            form.save()
            return redirect("expenses:home")
    return render(
        request, "expenses/set_budget.html", {"form": BudgetForm(instance=budget)}
    )


@login_required
def view_expenses(request):
    """
    View all expenses with filtering options.
    Admin users see all expenses, regular users see only their own.
    Supports filtering by category, date range, time period, and user (admin only).
    """
    # Get filter parameters
    filter_type = request.GET.get("filter", "all")  # all, this_month, today
    category_id = request.GET.get("category")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    user_id = request.GET.get("user")  # Admin can filter by user
    
    # Base queryset - admin sees all, users see only their own
    if request.user.is_superuser:
        expenses = Expense.objects.all().select_related("category", "member", "user").order_by("-date")
    else:
        expenses = Expense.objects.filter(user=request.user).select_related("category", "member").order_by("-date")
    
    # Apply user filter (admin only)
    if request.user.is_superuser and user_id:
        expenses = expenses.filter(user_id=user_id)
    
    # Apply time-based filters
    today = datetime.today()
    if filter_type == "this_month":
        expenses = expenses.filter(date__year=today.year, date__month=today.month)
    elif filter_type == "today":
        expenses = expenses.filter(date=today.date())
    
    # Apply category filter
    if category_id:
        expenses = expenses.filter(category_id=category_id)
    
    # Apply date range filters
    if from_date:
        expenses = expenses.filter(date__gte=from_date)
    if to_date:
        expenses = expenses.filter(date__lte=to_date)

    # Excel export
    if request.GET.get("export") == "1":
        wb = openpyxl.Workbook()
        ws = wb.active
        if request.user.is_superuser:
            ws.append(["Date", "User", "Category", "Description", "Amount"])
            for e in expenses:
                ws.append([
                    e.date.strftime("%d %b, %Y"),
                    e.user.username,
                    e.category.name if e.category else "General",
                    e.description or "",
                    float(e.amount),
                ])
        else:
            ws.append(["Date", "Category", "Description", "Amount"])
            for e in expenses:
                ws.append([
                    e.date.strftime("%d %b, %Y"),
                    e.category.name if e.category else "General",
                    e.description or "",
                    float(e.amount),
                ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=filtered_expenses.xlsx"
        wb.save(response)
        return response

    # Calculate totals
    total_amount = expenses.aggregate(Sum("amount"))["amount__sum"] or 0
    total_records = expenses.count()
    
    categories = ExpenseCategory.objects.all()
    
    # Get all users for admin filter
    from django.contrib.auth.models import User
    all_users = User.objects.all().order_by('username') if request.user.is_superuser else []
    
    context = {
        "expenses": expenses,
        "categories": categories,
        "total_amount": total_amount,
        "total_records": total_records,
        "filter_type": filter_type,
        "is_admin": request.user.is_superuser,
        "all_users": all_users,
    }
    
    return render(request, "expenses/view_expenses.html", context)


@login_required
def expense_stats(request):
    today = datetime.today()
    monthly_expenses = Expense.objects.filter(
        user=request.user, date__year=today.year, date__month=today.month
    )
    category_summary = (
        monthly_expenses.values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )
    context = {
        "labels": json.dumps(
            [item["category__name"] or "General" for item in category_summary]
        ),
        "data": json.dumps([float(item["total"]) for item in category_summary]),
        "category_summary": category_summary,
        "current_month": today.strftime("%B %Y"),
    }
    return render(request, "expenses/stats.html", context)


# --- User Management (Admin Only) ---
@login_required
def manage_users(request):
    """Admin can view and manage all users"""
    if not request.user.is_superuser:
        messages.error(request, "Access denied. Admin only.")
        return redirect("expenses:home")
    
    from django.contrib.auth.models import User
    users = User.objects.all().order_by('-date_joined')
    
    context = {
        'users': users,
    }
    return render(request, 'expenses/manage_users.html', context)


@login_required
def reset_user_password(request, user_id):
    """Admin can reset any user's password"""
    if not request.user.is_superuser:
        messages.error(request, "Access denied. Admin only.")
        return redirect("expenses:home")
    
    from django.contrib.auth.models import User
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if new_password and new_password == confirm_password:
            user.set_password(new_password)
            user.save()
            messages.success(request, f"Password reset successful for user: {user.username}")
            return redirect("expenses:manage_users")
        else:
            messages.error(request, "Passwords do not match!")
    
    context = {
        'user_to_reset': user,
    }
    return render(request, 'expenses/reset_password.html', context)
